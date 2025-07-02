"""
Excel Reader Module - Handles reading and parsing Excel files
"""

import pandas as pd
import os
from pathlib import Path
from typing import List, Dict, Optional, Any
import openpyxl
from openpyxl import load_workbook

from ..utils.helpers import DateParser, NumberParser

class ExcelReader:
    """Handles Excel file reading and data preprocessing"""
    
    def __init__(self, logger):
        self.logger = logger
        self.current_file = None
        self.workbook = None
        self.sheet_names = []
        
    def scan_files(self, directory_path: str) -> List[Dict[str, str]]:
        """
        Scan directory for Excel files
        
        Args:
            directory_path (str): Path to directory to scan
            
        Returns:
            List[Dict]: List of file information dictionaries
        """
        files = []
        
        try:
            directory = Path(directory_path)
            if not directory.exists():
                self.logger.warning(f"Directory does not exist: {directory_path}")
                return files
            
            # Look for Excel files
            excel_extensions = ['.xlsx', '.xls', '.xlsm']
            
            for file_path in directory.rglob('*'):
                if file_path.is_file() and file_path.suffix.lower() in excel_extensions:
                    # Skip temporary files
                    if file_path.name.startswith('~$'):
                        continue
                    
                    file_info = {
                        'name': file_path.name,
                        'path': str(file_path),
                        'size': file_path.stat().st_size,
                        'modified': file_path.stat().st_mtime
                    }
                    files.append(file_info)
            
            # Sort by modification time (newest first)
            files.sort(key=lambda x: x['modified'], reverse=True)
            
            self.logger.info(f"Found {len(files)} Excel files in {directory_path}")
            
        except Exception as e:
            self.logger.error(f"Error scanning directory {directory_path}: {str(e)}")
        
        return files
    
    def load_file(self, file_path: str) -> bool:
        """
        Load Excel file and get basic information
        
        Args:
            file_path (str): Path to Excel file
            
        Returns:
            bool: True if loaded successfully
        """
        try:
            self.current_file = file_path
            self.workbook = load_workbook(file_path, read_only=True, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            
            self.logger.info(f"Loaded Excel file: {file_path}")
            self.logger.info(f"Found sheets: {', '.join(self.sheet_names)}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading Excel file {file_path}: {str(e)}")
            self.current_file = None
            self.workbook = None
            self.sheet_names = []
            return False
    
    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names from loaded workbook"""
        return self.sheet_names.copy()
    
    def get_sheet_info(self, sheet_name: str) -> Dict[str, Any]:
        """
        Get information about a specific sheet
        
        Args:
            sheet_name (str): Name of the sheet
            
        Returns:
            Dict: Sheet information
        """
        try:
            if not self.workbook or sheet_name not in self.sheet_names:
                return {}
            
            sheet = self.workbook[sheet_name]
            
            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # Get first few rows to identify headers
            header_data = []
            for row in range(1, min(6, max_row + 1)):  # First 5 rows
                row_data = []
                for col in range(1, max_col + 1):  # Read ALL columns, not just first 20
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                header_data.append(row_data)
            
            # Get column headers (assume first row)
            columns = []
            if header_data:
                # Read ALL columns from the first row
                for i, header in enumerate(header_data[0]):
                    columns.append({
                        'index': i,
                        'name': header or f"Column {i + 1}",
                        'letter': openpyxl.utils.get_column_letter(i + 1)
                    })
            
            return {
                'name': sheet_name,
                'rows': max_row,
                'columns': max_col,
                'headers': header_data,
                'column_info': columns
            }
            
        except Exception as e:
            self.logger.error(f"Error getting sheet info for {sheet_name}: {str(e)}")
            return {}
    
    def read_data(self, sheet_name: str, column_mapping: Dict[str, str], 
                  date_format: str = "auto", number_format: str = "auto",
                  header_row: int = 1) -> pd.DataFrame:
        """
        Read and parse data from Excel sheet
        
        Args:
            sheet_name (str): Name of the sheet to read
            column_mapping (Dict[str, str]): Mapping of field names to column names
            date_format (str): Date format preference
            number_format (str): Number format preference
            header_row (int): Row number containing headers (1-based)
            
        Returns:
            pd.DataFrame: Processed data
        """
        try:
            if not self.current_file:
                raise ValueError("No file loaded")
            
            # Read with pandas for easier processing
            df = pd.read_excel(
                self.current_file,
                sheet_name=sheet_name,
                header=header_row - 1,  # pandas uses 0-based indexing
                engine='openpyxl'
            )
            
            self.logger.info(f"Read {len(df)} rows from sheet '{sheet_name}'")
            
            # Process column mapping
            processed_data = {}
            
            for field_name, column_name in column_mapping.items():
                if column_name and column_name in df.columns:
                    column_data = df[column_name].copy()
                    
                    # Apply specific processing based on field type
                    if field_name.lower() in ['date', 'invoice_date', 'shipment_date']:
                        processed_data[field_name] = self._process_date_column(
                            column_data, date_format
                        )
                    elif field_name.lower() in ['unit_price', 'quantity', 'total_value', 'price']:
                        processed_data[field_name] = self._process_number_column(
                            column_data, number_format
                        )
                    else:
                        processed_data[field_name] = self._process_text_column(column_data)
                else:
                    # Create empty column if not mapped
                    processed_data[field_name] = [None] * len(df)
            
            # Create processed DataFrame
            result_df = pd.DataFrame(processed_data)
            
            # Remove completely empty rows
            result_df = result_df.dropna(how='all')
            
            self.logger.info(f"Processed {len(result_df)} valid rows")
            
            return result_df
            
        except Exception as e:
            self.logger.error(f"Error reading data from {sheet_name}: {str(e)}")
            return pd.DataFrame()
    
    def _process_date_column(self, column: pd.Series, date_format: str) -> pd.Series:
        """Process date column with specified format"""
        def parse_date_value(value):
            return DateParser.parse_date(value, date_format)
        
        return column.apply(parse_date_value)
    
    def _process_number_column(self, column: pd.Series, number_format: str) -> pd.Series:
        """Process numeric column with specified format"""
        def parse_number_value(value):
            return NumberParser.parse_number(value, number_format)
        
        return column.apply(parse_number_value)
    
    def _process_text_column(self, column: pd.Series) -> pd.Series:
        """Process text column"""
        def clean_text_value(value):
            if pd.isna(value) or value is None:
                return None
            return str(value).strip()
        
        return column.apply(clean_text_value)
    
    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
        self.current_file = None
        self.sheet_names = []
