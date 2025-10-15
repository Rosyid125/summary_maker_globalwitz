"""
Output Formatter Module - Handles Excel output generation with formatting
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from typing import Dict, List, Any, Optional

from ..utils.helpers import format_currency, get_month_name, format_american_number, format_qty_with_precision

class OutputFormatter:
    """Handles Excel output generation with complex formatting"""
    
    def __init__(self, logger):
        self.logger = logger
        self.workbook = None
        self.quarter_colors = {
            1: PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),  # Light Red
            2: PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),  # Light Blue
            3: PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),  # Light Green
            4: PatternFill(start_color="FFFFE6", end_color="FFFFE6", fill_type="solid"),  # Light Yellow
        }
        
        # Define styles
        self.header_font = Font(bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Number format strings for Excel
        self.number_format_integer = '#,##0'  # For integers with comma separator
        self.number_format_decimal = '#,##0.000'  # For decimals with 3 decimal places
    
    def _write_numeric_cell(self, ws, row: int, col: int, value, decimals: int = 3, 
                           apply_border: bool = True, apply_fill=None):
        """
        Write a numeric value to cell with proper Excel number format
        
        Args:
            ws: Worksheet
            row: Row number
            col: Column number
            value: Numeric value (will be written as number, not string)
            decimals: Number of decimal places (0 for integer, 3 for decimal)
            apply_border: Whether to apply border
            apply_fill: Fill pattern to apply
        """
        cell = ws.cell(row=row, column=col)
        
        # Handle None or empty values
        if value is None or (isinstance(value, str) and value in ["-", ""]):
            cell.value = None
            if apply_border:
                cell.border = self.border_thin
            cell.alignment = Alignment(horizontal='right')
            return cell
        
        try:
            # Convert to float/int
            numeric_value = float(value)
            
            # Write numeric value (not string)
            cell.value = numeric_value
            
            # Set number format
            if decimals == 0 or numeric_value == int(numeric_value):
                cell.number_format = self.number_format_integer
            else:
                cell.number_format = self.number_format_decimal
            
            # Apply styling
            if apply_border:
                cell.border = self.border_thin
            if apply_fill:
                cell.fill = apply_fill
            
            cell.alignment = Alignment(horizontal='right')
            
        except (ValueError, TypeError):
            # If conversion fails, write as is
            cell.value = value
            if apply_border:
                cell.border = self.border_thin
        
        return cell
    
    def create_output_file(self, aggregated_data: Dict[str, Any], output_path: str, 
                          incoterm: str = "FOB", year: int = None) -> bool:
        """
        Create formatted Excel output file
        
        Args:
            aggregated_data (Dict): Aggregated data from DataAggregator
            output_path (str): Output file path
            incoterm (str): INCOTERM for pricing
            year (int): Year for the report
            
        Returns:
            bool: True if successful
        """
        try:
            self.workbook = Workbook()
            
            # Remove default sheet
            if 'Sheet' in self.workbook.sheetnames:
                self.workbook.remove(self.workbook['Sheet'])
            
            # Create sheets for each importer
            for importer_name, importer_data in aggregated_data.items():
                self._create_importer_sheet(importer_name, importer_data, incoterm, year)
            
            # Create summary sheet
            self._create_summary_sheet(aggregated_data, incoterm, year)
            
            # Save workbook
            self.workbook.save(output_path)
            self.logger.info(f"Output file created: {output_path}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating output file: {str(e)}")
            return False
    
    def _create_importer_sheet(self, importer_name: str, importer_data: Dict[str, Any], 
                              incoterm: str, year: int):
        """Create sheet for individual importer"""
        try:
            # Create worksheet
            sheet_name = self._sanitize_sheet_name(importer_name)
            ws = self.workbook.create_sheet(title=sheet_name)
            
            current_row = 1
            
            # Title
            ws.cell(row=current_row, column=1, value=f"Import Summary - {importer_name}")
            ws.cell(row=current_row, column=1).font = self.title_font
            current_row += 2
            
            # Overall summary
            if 'overall_summary' in importer_data:
                current_row = self._add_overall_summary(ws, current_row, importer_data['overall_summary'], incoterm)
                current_row += 2
            
            # Monthly summary
            if 'monthly_summary' in importer_data and importer_data['monthly_summary']:
                current_row = self._add_monthly_summary(ws, current_row, importer_data['monthly_summary'], incoterm)
                current_row += 2
            
            # Supplier summary
            if 'supplier_summary' in importer_data and importer_data['supplier_summary']:
                current_row = self._add_supplier_summary(ws, current_row, importer_data['supplier_summary'], incoterm)
                current_row += 2
            
            # Item summary
            if 'item_summary' in importer_data and importer_data['item_summary']:
                current_row = self._add_item_summary(ws, current_row, importer_data['item_summary'], incoterm)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            self.logger.error(f"Error creating sheet for {importer_name}: {str(e)}")
    
    def _add_overall_summary(self, ws, start_row: int, summary: Dict[str, Any], incoterm: str) -> int:
        """Add overall summary section"""
        current_row = start_row
        
        # Section title
        ws.cell(row=current_row, column=1, value="Overall Summary")
        ws.cell(row=current_row, column=1).font = self.header_font
        current_row += 1
        
        # Summary data with raw numeric values
        ws.cell(row=current_row, column=1, value="Total Records").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=summary.get('total_records', 0))
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Total Quantity").font = Font(bold=True)
        self._write_numeric_cell(ws, current_row, 2, summary.get('total_quantity', 0), decimals=3, apply_border=False)
        current_row += 1
        
        ws.cell(row=current_row, column=1, value=f"Average Unit Price ({incoterm})").font = Font(bold=True)
        avg_price = summary.get('avg_unit_price', 0)
        if avg_price:
            self._write_numeric_cell(ws, current_row, 2, avg_price, decimals=3, apply_border=False)
        else:
            ws.cell(row=current_row, column=2, value="-")
        current_row += 1
        
        ws.cell(row=current_row, column=1, value=f"Total Value ({incoterm})").font = Font(bold=True)
        self._write_numeric_cell(ws, current_row, 2, summary.get('total_value', 0), decimals=3, apply_border=False)
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Unique Suppliers").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=summary.get('unique_suppliers', 0))
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Unique Items").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=summary.get('unique_items', 0))
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Unique HS Codes").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=summary.get('unique_hs_codes', 0))
        current_row += 1
        
        return current_row
    
    def _add_monthly_summary(self, ws, start_row: int, monthly_data: List[Dict[str, Any]], incoterm: str) -> int:
        """Add monthly summary table"""
        current_row = start_row
        
        # Section title
        ws.cell(row=current_row, column=1, value="Monthly Summary")
        ws.cell(row=current_row, column=1).font = self.header_font
        current_row += 1
        
        # Headers
        headers = [
            "Month", "HS Code", "Item", "GSM", "Add On", 
            "Total Quantity", f"Avg Unit Price ({incoterm})", f"Total Value ({incoterm})", "Records"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.border = self.border_thin
            cell.alignment = self.center_alignment
        
        current_row += 1
        
        # Data rows
        for row_data in monthly_data:
            quarter = self._get_quarter_from_month_year(row_data.get('month_year', ''))
            quarter_fill = self.quarter_colors.get(quarter, None)
            
            # Text columns
            ws.cell(row=current_row, column=1, value=row_data.get('month_name', '')).border = self.border_thin
            ws.cell(row=current_row, column=2, value=row_data.get('hs_code', '')).border = self.border_thin
            ws.cell(row=current_row, column=3, value=row_data.get('item', '')).border = self.border_thin
            ws.cell(row=current_row, column=4, value=row_data.get('gsm', '')).border = self.border_thin
            ws.cell(row=current_row, column=5, value=row_data.get('add_on', '')).border = self.border_thin
            
            # Numeric columns with proper formatting
            self._write_numeric_cell(ws, current_row, 6, row_data.get('total_quantity', 0), decimals=3, apply_fill=quarter_fill)
            
            avg_price = row_data.get('avg_unit_price', 0)
            if avg_price:
                self._write_numeric_cell(ws, current_row, 7, avg_price, decimals=3, apply_fill=quarter_fill)
            else:
                cell = ws.cell(row=current_row, column=7, value=None)
                cell.border = self.border_thin
                if quarter_fill:
                    cell.fill = quarter_fill
            
            self._write_numeric_cell(ws, current_row, 8, row_data.get('total_value', 0), decimals=3, apply_fill=quarter_fill)
            
            cell = ws.cell(row=current_row, column=9, value=row_data.get('record_count', 0))
            cell.border = self.border_thin
            cell.alignment = Alignment(horizontal='right')
            if quarter_fill:
                cell.fill = quarter_fill
            
            current_row += 1
        
        return current_row
    
    def _add_supplier_summary(self, ws, start_row: int, supplier_data: List[Dict[str, Any]], incoterm: str) -> int:
        """Add supplier summary table"""
        current_row = start_row
        
        # Section title
        ws.cell(row=current_row, column=1, value="Supplier Summary")
        ws.cell(row=current_row, column=1).font = self.header_font
        current_row += 1
        
        # Headers
        headers = [
            "Supplier", "Total Quantity", f"Avg Unit Price ({incoterm})", 
            f"Total Value ({incoterm})", "Records", "Unique Items"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.border = self.border_thin
            cell.alignment = self.center_alignment
        
        current_row += 1
        
        # Data rows
        for row_data in supplier_data:
            # Text column
            ws.cell(row=current_row, column=1, value=row_data.get('supplier', '')).border = self.border_thin
            
            # Numeric columns with proper formatting
            self._write_numeric_cell(ws, current_row, 2, row_data.get('total_quantity', 0), decimals=3)
            
            avg_price = row_data.get('avg_unit_price', 0)
            if avg_price:
                self._write_numeric_cell(ws, current_row, 3, avg_price, decimals=3)
            else:
                cell = ws.cell(row=current_row, column=3, value=None)
                cell.border = self.border_thin
            
            self._write_numeric_cell(ws, current_row, 4, row_data.get('total_value', 0), decimals=3)
            
            cell = ws.cell(row=current_row, column=5, value=row_data.get('record_count', 0))
            cell.border = self.border_thin
            cell.alignment = Alignment(horizontal='right')
            
            cell = ws.cell(row=current_row, column=6, value=row_data.get('unique_items', 0))
            cell.border = self.border_thin
            cell.alignment = Alignment(horizontal='right')
            
            current_row += 1
        
        return current_row
    
    def _add_item_summary(self, ws, start_row: int, item_data: List[Dict[str, Any]], incoterm: str) -> int:
        """Add item summary table"""
        current_row = start_row
        
        # Section title
        ws.cell(row=current_row, column=1, value="Item Summary")
        ws.cell(row=current_row, column=1).font = self.header_font
        current_row += 1
        
        # Headers
        headers = [
            "Item", "Total Quantity", f"Avg Unit Price ({incoterm})", 
            f"Total Value ({incoterm})", "Records", "Unique Suppliers", "Avg GSM"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.border = self.border_thin
            cell.alignment = self.center_alignment
        
        current_row += 1
        
        # Data rows
        for row_data in item_data:
            # Text column
            ws.cell(row=current_row, column=1, value=row_data.get('item', '')).border = self.border_thin
            
            # Numeric columns with proper formatting
            self._write_numeric_cell(ws, current_row, 2, row_data.get('total_quantity', 0), decimals=3)
            
            avg_price = row_data.get('avg_unit_price', 0)
            if avg_price:
                self._write_numeric_cell(ws, current_row, 3, avg_price, decimals=3)
            else:
                cell = ws.cell(row=current_row, column=3, value=None)
                cell.border = self.border_thin
            
            self._write_numeric_cell(ws, current_row, 4, row_data.get('total_value', 0), decimals=3)
            
            cell = ws.cell(row=current_row, column=5, value=row_data.get('record_count', 0))
            cell.border = self.border_thin
            cell.alignment = Alignment(horizontal='right')
            
            cell = ws.cell(row=current_row, column=6, value=row_data.get('unique_suppliers', 0))
            cell.border = self.border_thin
            cell.alignment = Alignment(horizontal='right')
            
            avg_gsm = row_data.get('avg_gsm', 0)
            if avg_gsm:
                self._write_numeric_cell(ws, current_row, 7, avg_gsm, decimals=3)
            else:
                cell = ws.cell(row=current_row, column=7, value=None)
                cell.border = self.border_thin
            
            current_row += 1
        
        return current_row
    
    def _create_summary_sheet(self, aggregated_data: Dict[str, Any], incoterm: str, year: int):
        """Create overall summary sheet"""
        try:
            ws = self.workbook.create_sheet(title="Overall Summary", index=0)
            
            current_row = 1
            
            # Title
            ws.cell(row=current_row, column=1, value=f"Import Summary Report - {year or 'All Years'}")
            ws.cell(row=current_row, column=1).font = self.title_font
            current_row += 2
            
            # Summary by importer
            ws.cell(row=current_row, column=1, value="Summary by Importer")
            ws.cell(row=current_row, column=1).font = self.header_font
            current_row += 1
            
            # Headers
            headers = [
                "Importer", "Total Records", "Total Quantity", 
                f"Total Value ({incoterm})", "Unique Suppliers", "Unique Items"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.header_font
                cell.border = self.border_thin
                cell.alignment = self.center_alignment
            
            current_row += 1
            
            # Data rows
            total_records = 0
            total_quantity = 0
            total_value = 0
            
            for importer_name, importer_data in aggregated_data.items():
                overall = importer_data.get('overall_summary', {})
                
                # Update totals
                total_records += overall.get('total_records', 0)
                total_quantity += overall.get('total_quantity', 0)
                total_value += overall.get('total_value', 0)
                
                # Text column
                ws.cell(row=current_row, column=1, value=importer_name).border = self.border_thin
                
                # Numeric columns with proper formatting
                cell = ws.cell(row=current_row, column=2, value=overall.get('total_records', 0))
                cell.border = self.border_thin
                cell.alignment = Alignment(horizontal='right')
                
                self._write_numeric_cell(ws, current_row, 3, overall.get('total_quantity', 0), decimals=3)
                self._write_numeric_cell(ws, current_row, 4, overall.get('total_value', 0), decimals=3)
                
                cell = ws.cell(row=current_row, column=5, value=overall.get('unique_suppliers', 0))
                cell.border = self.border_thin
                cell.alignment = Alignment(horizontal='right')
                
                cell = ws.cell(row=current_row, column=6, value=overall.get('unique_items', 0))
                cell.border = self.border_thin
                cell.alignment = Alignment(horizontal='right')
                
                current_row += 1
            
            # Total row
            total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            cell = ws.cell(row=current_row, column=1, value="TOTAL")
            cell.font = Font(bold=True)
            cell.border = self.border_thin
            cell.fill = total_fill
            
            cell = ws.cell(row=current_row, column=2, value=total_records)
            cell.font = Font(bold=True)
            cell.border = self.border_thin
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal='right')
            
            cell = self._write_numeric_cell(ws, current_row, 3, total_quantity, decimals=3, apply_fill=total_fill)
            cell.font = Font(bold=True)
            
            cell = self._write_numeric_cell(ws, current_row, 4, total_value, decimals=3, apply_fill=total_fill)
            cell.font = Font(bold=True)
            
            # Can't sum unique suppliers/items across importers
            for col in [5, 6]:
                cell = ws.cell(row=current_row, column=col, value="")
                cell.font = Font(bold=True)
                cell.border = self.border_thin
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='right')
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            self.logger.error(f"Error creating summary sheet: {str(e)}")
    
    def _get_quarter_from_month_year(self, month_year: str) -> int:
        """Get quarter number from month_year string"""
        try:
            if not month_year or '-' not in month_year:
                return 1
            
            month = int(month_year.split('-')[1])
            return ((month - 1) // 3) + 1
            
        except Exception:
            return 1
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name for Excel compatibility"""
        if not name:
            return "Sheet1"
        
        # Remove invalid characters
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # Limit length
        return name[:31]
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # Set column width with some padding
                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            self.logger.error(f"Error adjusting column widths: {str(e)}")
